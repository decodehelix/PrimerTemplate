from nupack import *
import time, itertools
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import torch
import itertools

CheckMPS = torch.backends.mps.is_built()
device = "mps" if CheckMPS else "cuda" if torch.cuda.is_available() else "cpu"

Primers = ["ACCGAGATGATGTTTGCCCA",
"ATCTGTATCGCCCAGAGTAG",
"AGTCGGTCTATCATCAAGGC",
"AGGCCTTGAAATCCGACTGT",
"CGTCTCATAGGGCTGTCAAA",
"CACCGTTGATCAAGAGTTCG",
"CCGATTTGCAAGTACAGGTC",
"CCCAAAGTAGACGTTCGTGT",
"GTGATCCCAATGTCGCGTAA",
"GATAAGGCCTGCATACTCTG",
"GACCGAATTGTGACCATTGC",
"GTTGCGTATCCAGGACCAAT",
"TGCTTCAGGACCTAGTACGA",
"TATGTCCACCTGCGAGAATG",
"TAGATGCGAGTATGTCCACC",
"TGAGAACCTGCTGCATCGAT"]

def generate_primer_sequence(ForwardPrimer, writer, start_col, color):
    bases = ["A", "G", "C", "T"]
    print("Original primer Sequence: ", ForwardPrimer)
    ForwardPrimerReverseComplement = reverse_complement(ForwardPrimer)
    print("Reverse Complement Sequence: ", ForwardPrimerReverseComplement)
    for Temperature in Temperatures:
        PrimerModel = Model(material='dna', celsius=Temperature, sodium=0.5, magnesium=0.1)
        for n in range(0, 5):
            print(f"Calculation for {n} changes: ") 
            Primer_After_Changes = []
            sequence = ["".join(p) for p in itertools.product(bases, repeat=n)]
            primer = ForwardPrimer[0+n:len(ForwardPrimer)-n]
            for j in  range(len(sequence)):
                for k in  range(len(sequence)):
                    Primer_After_Changes.append(sequence[j] + primer + sequence[k])
            
            df = pd.DataFrame(Primer_After_Changes, columns=[f"Primers_After_{n}_Changes"])
                    
            dd = df[f"Primers_After_{n}_Changes"].tolist()
            FP_Energy = []
            FP_EnergyTemp = []
            Primer_Reverse_Complement = []
            Primer_Sequence = []
            dna_sequences = []
            Starting_Ending_Bases = []
            
            for l in range(len(dd)):  

                FP_New_String = dd[l] 
    
                PrimerRx = Strand(dd[l], name="PrimerRx")
                TemplateRx = Strand(ForwardPrimerReverseComplement, name="TemplateRx")
                t1 = Tube(strands={PrimerRx:5e-6, TemplateRx:5e-6}, complexes=SetSpec(max_size=2), name="Tube t1")
                tube_result_1 = tube_analysis(tubes=[t1], compute=["pairs", "mfe"], model=PrimerModel)
                dna_sequences.append(FP_New_String)
                Starting_Ending_Bases.append(FP_New_String[0] + FP_New_String[-1])
                try:
                    walker_result_1 = tube_result_1["(PrimerRx+TemplateRx)"]
                except KeyError:
                    walker_result_1 = tube_result_1["(TemplateRx+PrimerRx)"]
                for i, s in enumerate(walker_result_1.mfe):
                    Primer_Sequence.append(FP_New_String)
                    Primer_Reverse_Complement.append(ForwardPrimerReverseComplement[::-1])
                    FP_EnergyTemp.append(float("%.2f" % (s.energy)))
                
            FP_Energy = [x for x in FP_EnergyTemp if x != "None"]
    
            min_length = min(len(Primer_Sequence), len(Primer_Reverse_Complement), len(FP_Energy))
            max_length = max(len(Primer_Sequence), len(Primer_Reverse_Complement), len(FP_Energy))
            
            Primer_Sequence, Primer_Reverse_Complement, FP_Energy = [i + [None]*(max_length - len(i)) for i in [Primer_Sequence, Primer_Reverse_Complement, FP_Energy]]
            
            PrimerData = pd.DataFrame(   {"Primer Sequence"    : Primer_Sequence, 
                                         "Primer Reverse Complement"    : Primer_Reverse_Complement, 
                                         "FP Energy (kcal/mol)"     : FP_Energy}, 
                            columns =   [  "Primer Sequence", "Primer Reverse Complement" , "FP Energy (kcal/mol)"])
            
            sheet_name = f"Changes_{n} at {Temperature} \u00B0C"
            PrimerData.to_excel(writer, sheet_name=sheet_name, index=False, startcol=start_col, header=False)
            
            worksheet = writer.sheets[sheet_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=start_col+1, max_col=start_col+3):
                for cell in row:
                    cell.fill = fill
            print(n, Temperature)
writer = pd.ExcelWriter("Primers.xlsx", engine='openpyxl')

colors = [
    "F5B7B7", "FCD9D9", "F5E6B7", "F2F5B7", "D7F5B7", "B7F5DC", "C6F1EE", "C6E4F1",
    "C6D5F1", "CAC6F1", "DAC6F1", "EBC6F1", "F1C6EA", "F1C6D5", "F1C6C9", "F5D0B7"
]

start_col = 0
Temperatures = [35, 50, 65]
for i, Primer in enumerate(Primers):
    generate_primer_sequence(Primer, writer, start_col, colors[i % len(colors)])
    start_col += 3  
    
writer.close()
        

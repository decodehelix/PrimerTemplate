import pandas as pd
from nupack import *
import openpyxl
from openpyxl.styles import PatternFill
import torch

CheckMPS = torch.backends.mps.is_built()
device = "mps" if CheckMPS else "cuda" if torch.cuda.is_available() else "cpu"

Primers = ["ACCGAGATGATGTTTGCCCA",
"ATCTGTATCGCCCAGAGTAG",
"AGTCGGTCTATCATCAAGGC",
"AGGCCTTGAAATCCGACTGT",
"CGTCTCATAGGGCTGTCAAA",
"CACCGTTGATCAAGAGTTCG",
"CCGATTTGCAAGTACAGGTC",
"CCCAAAGTAGACGTTCGTGT",
"GTGATCCCAATGTCGCGTAA",
"GATAAGGCCTGCATACTCTG",
"GACCGAATTGTGACCATTGC",
"GTTGCGTATCCAGGACCAAT",
"TGCTTCAGGACCTAGTACGA",
"TATGTCCACCTGCGAGAATG",
"TAGATGCGAGTATGTCCACC",
"TGAGAACCTGCTGCATCGAT"]

def PrimerEnergyCalculator(Nucleotides, Temperatures):
    for Temperature in Temperatures:
        for Sodium in SodiumConcentration:
            for Magnesium in MagnesiumConcentration:
                ModifiedPrimerData = []  
                PrimerModel = Model(material='dna', celsius=Temperature, sodium=Sodium, magnesium=Magnesium)
        
                for index, Primer in enumerate(Primers):
                    PrimerEnergy = []  
                    PrimerReverseComplement = []  
                    PrimerSequence = []  
                    
                    PrimerReverseComplementSequence = reverse_complement(Primer)
                    for j in range(len(PrimerReverseComplementSequence)):
                        for Nucleotide in Nucleotides:
                            PrimerNewString = Primer[:j] + Nucleotide + Primer[j+1:]
                            
                            if PrimerNewString == Primer:
                                continue
                            
                            PrimerRx = Strand(PrimerNewString, name='PrimerRx')
                            TemplateRx = Strand(PrimerReverseComplementSequence, name='TemplateRx')
                            ComplexAnalysis = Tube(strands={PrimerRx:5e-6, TemplateRx:5e-6}, complexes=SetSpec(max_size=2), name='Tube ComplexAnalysis')
                            ComplexResults = tube_analysis(tubes=[ComplexAnalysis], compute=['pairs', 'mfe'], model=PrimerModel)
                            
                            try:
                                PrimerTemplateComplexResult = ComplexResults['(PrimerRx+TemplateRx)']
                            except KeyError:
                                PrimerTemplateComplexResult = ComplexResults['(TemplateRx+PrimerRx)']
                            
                            for FreeEnergy in PrimerTemplateComplexResult.mfe:
                                PrimerSequence.append(PrimerNewString)
                                PrimerReverseComplement.append(PrimerReverseComplementSequence[::-1])
                                PrimerEnergy.append(float('%.2f' % (FreeEnergy.energy)))
                    
                    ModifiedFreeEnergyData = pd.DataFrame({
                        f"Primer Sequence {index + 1}": PrimerSequence,
                        f"Primer Reverse Complement {index + 1}": PrimerReverseComplement,
                        f"Primer Energy (kcal/mol) {index + 1}": PrimerEnergy
                    })
                    
                    ModifiedPrimerData.append(ModifiedFreeEnergyData)
                
                ModifiedFreeEnergy = pd.concat(ModifiedPrimerData, axis=1)
                
                with pd.ExcelWriter("Mispriming.xlsx", engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                    SheetName = f"Modified at {Temperature} \u00B0C Na {Sodium} Mg {Magnesium}"
                    ModifiedFreeEnergy.to_excel(writer, sheet_name=SheetName, index=False)
                
                print(f"Modified Free Energy of primers saved in a sheet named {SheetName}.")

                workbook = openpyxl.load_workbook("Mispriming.xlsx")
                worksheet = workbook[SheetName]
                
                colors = [
                            "F5B7B7", "FCD9D9", "F5E6B7", "F2F5B7", "D7F5B7", "B7F5DC", "C6F1EE", "C6E4F1",
                            "C6D5F1", "CAC6F1", "DAC6F1", "EBC6F1", "F1C6EA", "F1C6D5", "F1C6C9", "F5D0B7"
                        ]

                for col_index in range(len(Primers)):
                    fill_color = PatternFill(start_color=colors[col_index], end_color=colors[col_index], fill_type="solid")
                    for row in worksheet.iter_rows(min_row=2, min_col=3*col_index+1, max_col=3*col_index+3):
                        for cell in row:
                            cell.fill = fill_color
                
                workbook.save("Mispriming.xlsx")

                PrimerReverseComplement = []  
                PrimerSequence = []
                PrimerEnergy = []
                for Primer in Primers:
                    
                    PrimerReverseComplementSequence = reverse_complement(Primer)
                    
                    PrimerRx = Strand(Primer, name='PrimerRx')
                    TemplateRx = Strand(PrimerReverseComplementSequence, name='TemplateRx')
                    ComplexAnalysis = Tube(strands={PrimerRx: 5e-6, TemplateRx: 5e-6}, complexes=SetSpec(max_size=2), name='Tube ComplexAnalysis')
                    ComplexResults = tube_analysis(tubes=[ComplexAnalysis], compute=['pairs', 'mfe'], model=PrimerModel)
                
                    try:
                        ComplexResults = tube_analysis(tubes=[ComplexAnalysis], compute=['pairs', 'mfe'], model=PrimerModel)
                        PrimerTemplateComplexResult = ComplexResults['(PrimerRx+TemplateRx)']
                    except (ValueError, KeyError):
                        PrimerTemplateComplexResult = ComplexResults['(TemplateRx+PrimerRx)']
                    
                    for FreeEnergy in PrimerTemplateComplexResult.mfe:
                        PrimerSequence.append(Primer)
                        PrimerReverseComplement.append(PrimerReverseComplementSequence[::-1])
                        PrimerEnergy.append(round(FreeEnergy.energy, 2))
                
                OriginalFreeEnergy = pd.DataFrame({
                    "Primer Sequence": PrimerSequence,
                    "Primer Reverse Complement ": PrimerReverseComplement,
                    "Primer Energy (kcal/mol)": PrimerEnergy
                })
                
                with pd.ExcelWriter("Mispriming.xlsx", engine='openpyxl', mode='a', if_sheet_exists="replace") as writer:
                    SheetName = f"Original at {Temperature} \u00B0C Na {Sodium} Mg {Magnesium}"
                    OriginalFreeEnergy.to_excel(writer, sheet_name=SheetName, index=False)
            
                print(f"Original Free Energy of primers saved in a sheet named {SheetName}.")

Nucleotides = ["A", "T", "G", "C"]
Temperatures = [35, 50, 65]
SodiumConcentration = [0.1, 0.5, 1.0]
MagnesiumConcentration = [0.0, 0.1, 0.2]

PrimerEnergyCalculator(Nucleotides, Temperatures)
